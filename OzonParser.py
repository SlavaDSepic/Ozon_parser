from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
import os
from selenium.webdriver.common.action_chains import ActionChains
from datetime import date


def Ozon_main():

    current_date = date.today()
    product_list = []

    os.environ['MOZ_HEADLESS'] = '1'


    options1 = Options()
    options1.add_argument('-headless')
    options1.binary_location = r"C:\Program Files\Firefox Developer " \
                               r"Edition\firefox.exe"
    options1.set_preference("general.useragent.override", UserAgent().random)
    options1.set_preference("network.websocket.enabled", False)
    options1.add_argument('window-size=1600x900')
    profile = webdriver.FirefoxProfile()
    profile.set_preference("dom.webdriver.enabled", False)
    profile.set_preference('useAutomationExtension', False)
    profile.set_preference('browser.migration.version', 9001)
    profile.update_preferences()

    driver = webdriver.Firefox(options=options1,
                               executable_path=r'C:\geckodriver' \
                                               r'.exe',
                               firefox_profile=profile)
    driver.implicitly_wait(5)

    view_port_height = "var viewPortHeight = Math.max(document.documentElement.clientHeight, window.innerHeight || 0);"
    element_top = "var elementTop = arguments[0].getBoundingClientRect().top;"
    js_function = "window.scrollBy(0, elementTop-(viewPortHeight/2));"
    scroll_into_middle = view_port_height + element_top + js_function

    actions = ActionChains(driver)


    def get_start_page():


        driver.get("https://www.ozon.ru/")
        print('Подключение к сайту Ozon')
        driver.maximize_window()
        sleep(3)


    def change_region():

        region_button = driver.find_element_by_css_selector('._3MPl > div:nth-child(2)')
        region_button.click()
        region_name = driver.find_element_by_class_name('a7')
        region_name.click()
        sleep(2)


    def get_articles_list(in_list):

        with open(in_list, encoding='utf-8') as file:
            lines = file.readlines()
            arts_list = []
            for line in lines:
                line = line.strip()
                arts_list.append(line)
            return arts_list


    def get_products(article):

        print(f'Поиск товаров по артикулу {article}')
        try:
            clear_tb = driver.find_element_by_class_name('b7i7')
            clear_tb.click()
            #print('Поле поиска очищено')
            sleep(1)
        except Exception:
            pass
        input_tb = driver.find_element_by_class_name('b7i5')
        input_tb.clear()
        input_tb.send_keys(article)
        input_tb.send_keys(Keys.ENTER)
        sleep(2)


        def get_page_html():

            sleep(1)
            page_html = driver.page_source
            page_html = page_html
            page_html = str(page_html)
            all_html = page_html
            return all_html


        def get_page_data(page_html):

            soup = BeautifulSoup(page_html, 'lxml')
            all_products_tags = soup.find_all('div', class_="a0c6")
            print(f'Найдено {len(all_products_tags)} товаров')
            return all_products_tags


        def get_products_data(all_products_tags):

            for product_tag in all_products_tags:
                #  Цена
                try:
                    cost = product_tag.find('span', class_='b5v6 b5v7 c4v8')
                    cost = cost.text
                except Exception:
                    cost = product_tag.find('span', class_='b5v6 b5v7')
                    cost = cost.text

                #  Название
                try:
                    product_name = product_tag.find('span', class_='j4 as3 az a0f2 f-tsBodyL item b3u9')
                    product_name = product_name.text
                except Exception:
                    product_name = 'Нет названия'
                #  Ссылка на товар
                try:
                    product_href = product_tag.find('a', class_='tile-hover-target b3u9').get('href')
                    product_href = 'https://www.ozon.ru' + product_href
                except Exception:
                    product_href = 'Ссылка отсутствует'


                #  Продавец
                try:
                    seller_name = product_tag.find(lambda tag: tag.name == 'span' and 'Ozon,' in tag.text)
                    seller_name = seller_name.text
                    seller_name = seller_name.split('продавец ')
                    seller_name = seller_name[-1]
                except Exception:
                    seller_name = 'Продавец не указан'


                product_list.append([article, product_name, product_href, seller_name, cost])
                print(product_list[-1])

        try:
            driver.implicitly_wait(1)
            not_result = driver.find_element_by_class_name('b6q3')
            print(f"Товары по артиклю {article} не найдены")
            product_list.append([article, '-', '-', '-', '-'])
            driver.implicitly_wait(5)
        except Exception:

            get_products_data(get_page_data(get_page_html()))

            def next_pages():

                while True:

                    try:
                        next_page_button = driver.find_element_by_class_name('Дальше')
                        next_page_button.click()
                        get_products_data(get_page_data(get_page_html()))
                        sleep(1)
                    except Exception:
                        break

            next_pages()

        return product_list


    def excel_table(product_list):

        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'] = 'Артикул'
        sheet['A1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")
        sheet['B1'] = 'Название'
        sheet['B1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")
        sheet['C1'] = 'Ссылка на сайт Озона'
        sheet['C1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")
        sheet['D1'] = 'Поставщик'
        sheet['D1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")
        sheet['E1'] = 'Цена'
        sheet['E1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C",
                                       fill_type="solid")

        wrap_alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        wrap_alignment2 = Alignment(wrap_text=True, horizontal='left',
                                    vertical='center')

        sheet.cell(row=1, column=1).alignment = wrap_alignment2
        sheet.cell(row=1, column=2).alignment = wrap_alignment2
        sheet.cell(row=1, column=3).alignment = wrap_alignment2
        sheet.cell(row=1, column=4).alignment = wrap_alignment2
        sheet.cell(row=1, column=5).alignment = wrap_alignment2

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 10

        row_number = 2

        for product in product_list:
            sheet.cell(row=row_number, column=1).value = product[0]
            sheet.cell(row=row_number, column=1).alignment = wrap_alignment2
            sheet.cell(row=row_number, column=2).value = product[1]
            sheet.cell(row=row_number, column=2).alignment = wrap_alignment2
            sheet.cell(row=row_number, column=3).hyperlink = product[2]
            sheet.cell(row=row_number, column=3).value = 'Ссылка на товар'
            sheet.cell(row=row_number, column=3).style = 'Hyperlink'
            #sheet.cell(row=row_number, column=3).value = '=HYPERLINK("{}", "{}")'.format(product[2], "Ссылка на товар")
            sheet.cell(row=row_number, column=3).alignment = wrap_alignment2
            sheet.cell(row=row_number, column=4).value = product[3]
            sheet.cell(row=row_number, column=4).alignment = wrap_alignment2
            sheet.cell(row=row_number, column=5).value = product[4]
            sheet.cell(row=row_number, column=5).alignment = wrap_alignment2
            row_number += 1


        c = sheet['A2']
        sheet.freeze_panes = c
        book_name = f'Прайс лист {current_date}.xlsx'
        book.save(book_name)
        print('Все данные с сайта Ozon собраны')
        book.close()

    get_start_page()
    # change_region()
    sleep(1)
    arts_list = get_articles_list('List.txt')
    for article in arts_list:
        try:
            get_products(article)
            excel_table(product_list)
        except Exception:
            try:
                driver.refresh()
                get_products(article)
                excel_table(product_list)
            except Exception:
                with open(f'Not_founds {current_date}.txt', 'a') as nfile:
                    nfile.write(article + '\n')
                    print(f'При поиске {article} произошла ошибка')


    driver.close()

if __name__ == "__main__":
    Ozon_main()
